from openpyxl import load_workbook
import constant

def excelutil():
    workbook = load_workbook(filename=constant.CaseFile)
    sheet = workbook.active

    datas = []
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):

            data = {"caseid": row[0],
                    "casename": row[1],
                    "zipdir": row[2],
                    "innerzipdir": row[3],
                    "filename": row[4],
                    "key": row[5],
                    "expected": str(row[6]),
                    "category": row[7],
                    "filesource": row[8]}
            datas.append(data)
        return datas
    except Exception as e:
        print("Error message is: %s" % e)

# datas = excelutil()
# print(excelutil()[1])
# print(excelutil())
# print(datas)
# print(len(datas))
# print(datas[0])



